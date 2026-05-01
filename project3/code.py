# Project 3: Auto Chart Generator
# Day 12 - Jessie Portfolio
# Generate chart + embed directly into Excel report

import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image


data = {
    'Month': ['Jan', 'Feb', 'Mar', 'Apr', 'May'],
    'Sales': [105000, 120000, 80000, 150000, 130000]
}
df = pd.DataFrame(data)

# 1. Create chart
plt.figure(figsize=(8,5))
plt.plot(df['Month'], df['Sales'], marker='o', color='#2E86C1', linewidth=3, markersize=8)

# Add data labels
for i, val in enumerate(df['Sales']):
    plt.text(i, val + 3000, f'₱{val:,.0f}', ha='center', fontsize=9)

plt.title('Monthly Sales Trend Report', fontsize=14, fontweight='bold')
plt.xlabel('Month', fontsize=12)
plt.ylabel('Sales (PHP)', fontsize=12)
plt.grid(True, linestyle='--', alpha=0.7)
plt.ylim(70000, 160000)
plt.tight_layout()

# Save as PNG for portfolio
chart_image = 'chart.jpg'
plt.savefig(chart_image, dpi=300, bbox_inches='tight')
plt.close()
print("✅ Chart image saved as chart.png")

# 2. Create Excel + embed chart
wb = Workbook()
ws = wb.active
ws.title = 'Sales Report'

# Add headers
ws['A1'] = 'Month'
ws['B1'] = 'Sales'
ws['A1'].font = ws['B1'].font = ws['A1'].font.copy(bold=True)

# Add data
for i, row in df.iterrows():
    ws.append([row['Month'], row['Sales']])
    ws[f'B{i+2}'].number_format = '₱#,##0'

# Add chart image to Excel
img = Image(chart_image)
img.width = 500
img.height = 300
ws.add_image(img, 'D2')

# Adjust column width
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 18

filename = "SALES_CHART.xlsx"
wb.save(filename)
print(f"✅ Excel file saved as {filename}")
print("Client gets 1 file with data + visual chart. No manual work needed.")
